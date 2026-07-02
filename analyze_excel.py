import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('/Users/a1708/Desktop/hxx/hxxPrivate/02结婚相关/4-9月血糖记录.xlsx')
sheet = wb.active

food_counter = Counter()
exercise_counter = Counter()

for row in range(3, sheet.max_row + 1):
    breakfast = str(sheet.cell(row=row, column=4).value or '').strip()
    lunch = str(sheet.cell(row=row, column=12).value or '').strip()
    dinner = str(sheet.cell(row=row, column=20).value or '').strip()
    
    for food in [breakfast, lunch, dinner]:
        if food:
            items = [f.strip() for f in food.split('\n') if f.strip()]
            for item in items:
                food_counter[item] += 1
    
    breakfast_ex = str(sheet.cell(row=row, column=5).value or '').strip()
    lunch_ex = str(sheet.cell(row=row, column=13).value or '').strip()
    dinner_ex = str(sheet.cell(row=row, column=21).value or '').strip()
    
    for ex in [breakfast_ex, lunch_ex, dinner_ex]:
        if ex:
            exercise_counter[ex] += 1

print('=== 餐食内容频率统计 ===')
for food, count in food_counter.most_common(30):
    print(f'{count:3d}次: {food}')

print()
print('=== 运动内容频率统计 ===')
for ex, count in exercise_counter.most_common(20):
    print(f'{count:3d}次: {ex}')
